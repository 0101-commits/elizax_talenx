# -*- coding: utf-8 -*-
"""신호 카탈로그 v0.6 JSON + W5 xlsx(문구 4건 갱신) -> js/ez_signals.js (window.EZSignalCatalog)

원천:
  C:\\Users\\cgpar\\elizax_docs\\성과관리Agent_신호카탈로그_v0.6.json   (신호 150건, 기계 단일 원천)
  ...05 HR AX\\00 제출본\\...W5_signal catalogue_260803.xlsx            (제출본. 알림 문구 4건이 더 다듬어짐)
산출:
  C:\\Users\\cgpar\\elizax_talenx\\js\\ez_signals.js
멱등. 다시 돌려도 같은 결과.
"""
import io
import json
import os
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

SRC_JSON = r"C:\Users\cgpar\elizax_docs\성과관리Agent_신호카탈로그_v0.6.json"
SRC_XLSX = (
    r"C:\Users\cgpar\OneDrive - 휴먼컨설팅그룹\03 Chapter\05 HR AX\00 제출본"
    r"\성과평가 AI Agent_AX Consulting_W5_signal catalogue_260803.xlsx"
)
OUT_JS = r"C:\Users\cgpar\elizax_talenx\js\ez_signals.js"

# 수신 대상 -> elizax 역할 키. 상위조직장은 별도 롤이 없어 조직장(leader)이 상위 관점으로 받는다.
ACTOR_ROLE = {
    "구성원": ["member"],
    "팀장": ["leader"],
    "상위조직장": ["leader"],
    "HR경영진": ["hr", "exec"],
}
ACTOR_ORDER = {"구성원": 1, "팀장": 2, "상위조직장": 3, "HR경영진": 4}

TYPE_LABEL = {
    "T1": "기한 도래",
    "T2": "작성 공백",
    "T3": "기준 이탈",
    "T4": "연결 불일치",
    "T5": "상황 변동",
}
ACTION_LABEL = {
    "A1": "새로 쓰기",
    "A2": "내가 고치기",
    "A3": "알려주기",
    "A4": "1on1 잡기",
    "A5": "상세 보기",
    "A6": "승인 요청",
}
STAGE_ORDER = {"목표수립": 1, "중간점검": 2, "평가": 3, "피드백": 4}


def load():
    sigs = json.load(open(SRC_JSON, encoding="utf-8"))
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb["신호 카탈로그"]
    # 열은 이름으로 찾는다 — 20차에 「이 알림을 부르는 질문」 열이 끼면서 뒤 열이 한 칸씩 밀렸다
    head = [(c or "").replace("\n", "") if isinstance(c, str) else "" for c in
            next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    def col(name):
        for i, h in enumerate(head):
            if h.startswith(name):
                return i
        raise SystemExit("열을 못 찾음: " + name)
    C_ID, C_AI, C_NOTICE, C_NEED = col("신호 ID"), col("AI 판단 가능"), col("알림(Signal) 문구"), col("필요 데이터")
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    assert len(rows) == len(sigs) == 150, (len(rows), len(sigs))
    fixed = 0
    for r, s in zip(rows, sigs):
        assert r[C_ID] == s["id"], (r[C_ID], s["id"])
        notice = (r[C_NOTICE] or "").strip()
        if notice and notice != s["notice"]:
            s["notice"] = notice          # 제출본 문구가 최신
            fixed += 1
        s["_ai"] = str(r[C_AI] or "").replace("\n", " ").strip()   # AI 판단 가능
        s["_need"] = str(r[C_NEED] or "").strip()                  # 필요 데이터
    print("notice synced from xlsx:", fixed)
    return sigs


def nowable(s):
    """새로 만들 기록 없이 지금 켤 수 있는 신호."""
    if s.get("todo_create"):
        return False
    return not any(a.get("newdata") for a in s["actions"])


def slim(s):
    ev = []
    for e in s["evidence"]:
        o = {
            "axis": e["axis"][1:] if e["axis"][:1] in "①②③④⑤⓪" else e["axis"],
            "mark": e["axis"][:1],
            "text": e["text"],
            "show": e["show"],
        }
        if e.get("emph"):
            o["emph"] = e["emph"]
        if e.get("src"):
            o["src"] = e["src"]
        if e.get("asof"):
            o["asof"] = e["asof"]
        if e.get("calc"):
            o["calc"] = e["calc"]
        if e.get("assumed"):
            o["assumed"] = 1
        if e.get("basis"):
            o["basis"] = e["basis"]
        ev.append(o)

    acts = []
    for a in s["actions"]:
        o = {
            "rank": a["rank"],
            "type": a["type"],
            "kind": ACTION_LABEL[a["type"]],
            "label": a["label"],
            "draft": a.get("draft", ""),
            "confirm": a.get("confirm", ""),
            "store": a.get("store", ""),
        }
        if a.get("chips"):
            o["chips"] = a["chips"]
        if a.get("newdata"):
            o["newdata"] = a["newdata"]
        acts.append(o)

    r = s["rule"]
    ths = [
        {"code": t["code"], "name": t["name"], "value": t["value"], "range": t.get("range", ""),
         "why": t.get("rationale", "")}
        for t in r.get("thresholds", [])
    ]
    out = {
        "id": s["id"],
        "no": s["no"],
        "stage": s["stage"],
        "stageNo": STAGE_ORDER[s["stage"]],
        "actor": s["actor"],
        "actorNo": ACTOR_ORDER[s["actor"]],
        "roles": ACTOR_ROLE[s["actor"]],
        "type": s["type"],
        "typeLabel": TYPE_LABEL[s["type"]],
        "level": s["level"],
        "ai": s["_ai"],
        "need": s["_need"],
        "now": 1 if nowable(s) else 0,
        "notice": s["notice"],
        "principle": r["principle"],
        "example": r["example"],
        "refs": r["refs"],
        "drafts": r["drafts"],
        "compare": r["compare"],
        "thresholds": ths,
        "evidence": ev,
        "actions": acts,
        "agent": s["agent0"],
        "done": s["done"],
        "score": s["score"],
        "mute": s["mute"],
        "adopt": s["adopt"],
    }
    if s.get("todo_decide"):
        out["todoDecide"] = s["todo_decide"]
    if s.get("todo_create"):
        out["todoCreate"] = s["todo_create"]
    return out


def main():
    sigs = load()
    cards = [slim(s) for s in sigs]
    payload = {
        "version": "v0.6 / W5 260803",
        "source": "성과평가 AI Agent_AX Consulting_W5_signal catalogue_260803.xlsx",
        "count": len(cards),
        "typeLabel": TYPE_LABEL,
        "actionLabel": ACTION_LABEL,
        "signals": cards,
    }
    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    js = (
        "/* ez_signals.js — 성과관리 Agent 신호 카탈로그 (자동 생성물, 직접 고치지 않는다)\n"
        "   생성기 : scripts/build_signals.py\n"
        "   원천   : 성과평가 AI Agent_AX Consulting_W5_signal catalogue_260803.xlsx (+ v0.6 JSON)\n"
        "   내용   : 신호 150건 · 근거/기준값/처리 방법/알림 문구. window.EZSignalCatalog 로 노출한다. */\n"
        "window.EZSignalCatalog = " + body + ";\n"
    )
    os.makedirs(os.path.dirname(OUT_JS), exist_ok=True)
    with io.open(OUT_JS, "w", encoding="utf-8", newline="\n") as f:
        f.write(js)
    kb = len(js.encode("utf-8")) / 1024
    now = [c for c in cards if c["now"]]
    print("wrote", OUT_JS, "%.0f KB" % kb)
    print("signals", len(cards), "| 지금 켤 수 있는 신호", len(now))
    from collections import Counter
    print("roles", dict(Counter(r for c in cards for r in c["roles"])))
    print("now by role", dict(Counter(r for c in now for r in c["roles"])))


if __name__ == "__main__":
    main()
